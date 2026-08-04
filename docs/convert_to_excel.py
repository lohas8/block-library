#!/usr/bin/env python3
"""将测试用例表转换为 Excel 格式"""

import re

# 读取测试用例表
with open('/root/.openclaw/workspace/block-library/docs/test-cases-table.md', 'r', encoding='utf-8') as f:
    content = f.read()

# 解析各模块
lines = content.split('\n')
module_data = []
current_module = None
current_table = []
in_table = False

for line in lines:
    # 检测模块标题 - 支持中文数字编号和阿拉伯数字编号
    if line.startswith('## '):
        if current_module and current_table:
            module_data.append((current_module, current_table))
        # 提取模块名和前缀 - 匹配各种编号格式
        match = re.match(r'## [一二三四五六七八九十\d]+[。.、].*?\((.+?)\)\s+特征前缀[：:]\s*(\w+)', line)
        if match:
            current_module = {
                'name': match.group(1).replace(' Module', ''),
                'prefix': match.group(2)
            }
            current_table = []
        in_table = False
    # 检测表头 - 跳过使用说明部分
    elif line.startswith('| 特征编号') or line.startswith('| 编号'):
        # 如果还没找到模块，说明这是使用说明表格，跳过
        if current_module is None:
            continue
        in_table = True
        continue
    # 检测分隔线
    elif re.match(r'\|[\s\-:]+\|', line):
        continue
    # 检测下一个模块或非数据行，重置in_table
    elif line.startswith('## ') or line.startswith('# ') or line.startswith('**'):
        in_table = False
    # 数据行
    elif line.startswith('|') and in_table and current_module:
        parts = [p.strip() for p in line.split('|')[1:-1]]
        # 检查是否是有效数据行 - 特征编号以字母-数字结尾
        if len(parts) >= 5 and parts[0] and parts[0].strip() != '':
            if re.match(r'^[A-Z]+-\d+$', parts[0].strip()):
                current_table.append(parts)

if current_module and current_table:
    module_data.append((current_module, current_table))

# 创建 Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "测试用例总表"

# 定义样式
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
p0_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # 浅红
p1_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 浅黄
p2_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # 浅绿

# 表头
headers = ["特征编号", "Level", "测试点", "测试步骤", "预期结果", "是否通过", "通过截图", "责任人", "测试时间", "备注"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# 填充数据
row = 2
total_cases = 0
for module_info, cases in module_data:
    for case in cases:
        if len(case) >= 5:
            total_cases += 1
            case_id = case[0]
            level = case[1]
            test_point = case[2]
            test_steps = case[3]
            expected = case[4]
            responsible = case[7] if len(case) > 7 else ""
            test_time = case[8] if len(case) > 8 else ""
            notes = case[9] if len(case) > 9 else ""
            
            ws.cell(row=row, column=1, value=case_id)
            ws.cell(row=row, column=2, value=level)
            ws.cell(row=row, column=3, value=test_point)
            ws.cell(row=row, column=4, value=test_steps)
            ws.cell(row=row, column=5, value=expected)
            ws.cell(row=row, column=6, value="待测试")
            ws.cell(row=row, column=7, value="")
            ws.cell(row=row, column=8, value=responsible)
            ws.cell(row=row, column=9, value=test_time)
            ws.cell(row=row, column=10, value=notes)
            
            # 根据Level设置行背景色
            fill = None
            if level == "P0":
                fill = p0_fill
            elif level == "P1":
                fill = p1_fill
            elif level == "P2":
                fill = p2_fill
            
            if fill:
                for col in range(1, 11):
                    ws.cell(row=row, column=col).fill = fill
            
            # 设置边框和对齐
            for col in range(1, 11):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            row += 1

# 设置列宽
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 28
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 35
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 12
ws.column_dimensions['J'].width = 15

# 冻结首行
ws.freeze_panes = 'A2'

# 添加统计sheet
ws_stats = wb.create_sheet("统计汇总")
stats_headers = ["模块", "P0用例数", "P1用例数", "P2用例数", "合计", "通过数", "通过率"]
for col, h in enumerate(stats_headers, 1):
    cell = ws_stats.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# 统计数据
stat_row = 2
total_p0 = total_p1 = total_p2 = 0
for module_info, cases in module_data:
    p0 = p1 = p2 = 0
    for case in cases:
        if len(case) >= 2:
            level = case[1]
            if level == "P0": p0 += 1
            elif level == "P1": p1 += 1
            elif level == "P2": p2 += 1
    
    total = p0 + p1 + p2
    total_p0 += p0
    total_p1 += p1
    total_p2 += p2
    
    ws_stats.cell(row=stat_row, column=1, value=module_info['name'])
    ws_stats.cell(row=stat_row, column=2, value=p0)
    ws_stats.cell(row=stat_row, column=3, value=p1)
    ws_stats.cell(row=stat_row, column=4, value=p2)
    ws_stats.cell(row=stat_row, column=5, value=total)
    ws_stats.cell(row=stat_row, column=6, value=0)
    ws_stats.cell(row=stat_row, column=7, value="0%")
    for col in range(1, 8):
        ws_stats.cell(row=stat_row, column=col).border = thin_border
    stat_row += 1

# 总计行
ws_stats.cell(row=stat_row, column=1, value="总计")
ws_stats.cell(row=stat_row, column=2, value=total_p0)
ws_stats.cell(row=stat_row, column=3, value=total_p1)
ws_stats.cell(row=stat_row, column=4, value=total_p2)
ws_stats.cell(row=stat_row, column=5, value=total_p0 + total_p1 + total_p2)
ws_stats.cell(row=stat_row, column=6, value=0)
ws_stats.cell(row=stat_row, column=7, value="0%")
for col in range(1, 8):
    ws_stats.cell(row=stat_row, column=col).border = thin_border
    ws_stats.cell(row=stat_row, column=col).font = Font(bold=True)

ws_stats.column_dimensions['A'].width = 20
ws_stats.column_dimensions['B'].width = 12
ws_stats.column_dimensions['C'].width = 12
ws_stats.column_dimensions['D'].width = 12
ws_stats.column_dimensions['E'].width = 10
ws_stats.column_dimensions['F'].width = 10
ws_stats.column_dimensions['G'].width = 10

# 保存
output_path = '/root/.openclaw/workspace/block-library/docs/测试用例表.xlsx'
wb.save(output_path)
print(f"Excel文件已生成: {output_path}")
print(f"共 {total_cases} 个测试用例")
print(f"模块数量: {len(module_data)}")
for m, c in module_data:
    print(f"  - {m['name']} ({m['prefix']}): {len(c)} 用例")
